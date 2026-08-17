"""Excel 表格比对核心逻辑。

流程：解析基准表与比对表的指定 sheet → 按键列归一化匹配 → 对选定值列逐一比较 →
生成差异明细表（独立 xlsx）或在原表格副本中高亮差异单元格（xlsx）。
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..excel_split.service import load_sheets
from .models import CompareRequest

# 差异类型标签（用于差异明细表与预览）
TYPE_LABELS = {
    "changed": "值差异",
    "only_base": "基准独有",
    "only_compare": "比对独有",
}

# 高亮颜色
FILL_CHANGED = PatternFill(fill_type="solid", fgColor="FFFF00")      # 黄：值差异单元格
FILL_ONLY_BASE = PatternFill(fill_type="solid", fgColor="F4B084")    # 橙：仅基准有
FILL_ONLY_COMPARE = PatternFill(fill_type="solid", fgColor="A9D08E")  # 绿：仅比对有
FILL_DIFF_TYPE = {
    "changed": FILL_CHANGED,
    "only_base": FILL_ONLY_BASE,
    "only_compare": FILL_ONLY_COMPARE,
}

PREVIEW_LIMIT = 100
MAX_ROW_STRING = 30000  # 整行拼接摘要的最大长度保护


class ExcelDiffError(ValueError):
    pass


# ---------------------------------------------------------------- 归一化

def normalize_cell(value: Any) -> str:
    """把任意单元格值归一化为可比较的字符串：空值统一为 ''，数值去掉无意义小数点，
    日期统一为 YYYY-MM-DD HH:MM:SS。"""
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, pd.Timedelta):
        return str(value)
    return str(value)


def _row_to_string(row: Dict[str, str]) -> str:
    """把一行数据拼成 '列=值；列=值' 的摘要，用于独有行的展示。"""
    parts = [f"{c}={v}" for c, v in row.items() if v != ""]
    text = "；".join(parts)
    return text[:MAX_ROW_STRING]


# ---------------------------------------------------------------- 解析

def rows_from_df(df: pd.DataFrame, key_col: str) -> Tuple[Dict[str, Dict[str, str]], List[str]]:
    """从 DataFrame（已 dropna 全空行）构建 key -> {非键列: 归一化值}，返回 (行, 表头)。"""
    headers = [str(c) for c in df.columns]
    if key_col not in headers:
        raise ExcelDiffError(f"表中不存在键列「{key_col}」")
    cols = [c for c in headers if c != key_col]
    result: Dict[str, Dict[str, str]] = {}
    seen: Dict[str, int] = {}
    for _, row in df.iterrows():
        key = normalize_cell(row[key_col])
        seen[key] = seen.get(key, 0) + 1
        result[key] = {c: normalize_cell(row[c]) for c in cols}
    _reject_duplicates(seen, key_col)
    return result, headers


def rows_from_sheet(
    ws, key_col: str, header_row: int = 1
) -> Tuple[Dict[str, Dict[str, str]], Dict[str, int], List[str]]:
    """从 openpyxl sheet 构建 key -> {非键列: 归一化值}，同时返回 key -> 行号映射（高亮用）。"""
    headers: List[str] = []
    for cell in ws[header_row]:
        headers.append("" if cell.value is None else str(cell.value))
    if key_col not in headers:
        raise ExcelDiffError(f"表中不存在键列「{key_col}」")
    key_idx = headers.index(key_col)
    cols = [c for c in headers if c != key_col]

    result: Dict[str, Dict[str, str]] = {}
    key_to_row: Dict[str, int] = {}
    seen: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=i).value for i in range(1, len(headers) + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue  # 跳过全空行
        key = normalize_cell(vals[key_idx])
        seen[key] = seen.get(key, 0) + 1
        result[key] = {headers[i]: normalize_cell(vals[i]) for i in range(len(headers)) if i != key_idx and i < len(vals)}
        key_to_row[key] = r
    _reject_duplicates(seen, key_col)
    return result, key_to_row, headers


def _reject_duplicates(seen: Dict[str, int], key_label: str) -> None:
    dup = [k for k, n in seen.items() if n > 1]
    if dup:
        sample = "、".join(dup[:5])
        raise ExcelDiffError(
            f"键列「{key_label}」存在重复值（如：{sample} 等共 {len(dup)} 个），"
            "无法唯一匹配两表，请检查数据后重试"
        )


# ---------------------------------------------------------------- 比对

def compute_diff(
    base_rows: Dict[str, Dict[str, str]],
    compare_rows: Dict[str, Dict[str, str]],
    key_label: str,
    value_cols: List[str],
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """按键匹配并逐值列比较，返回 (差异行列表, 统计)。"""
    base_keys = set(base_rows)
    compare_keys = set(compare_rows)
    matched_keys = base_keys & compare_keys

    diff_rows: List[Dict[str, Any]] = []
    changed = 0

    for key in sorted(matched_keys):
        changes = []
        for col in value_cols:
            bv = base_rows[key].get(col, "")
            cv = compare_rows[key].get(col, "")
            if bv != cv:
                changes.append((col, bv, cv))
        if changes:
            changed += 1
            for col, bv, cv in changes:
                diff_rows.append(
                    {
                        "diff_type": "changed",
                        "key_column": key_label,
                        "key_value": key,
                        "column": col,
                        "base_value": bv,
                        "compare_value": cv,
                    }
                )

    for key in sorted(base_keys - compare_keys):
        diff_rows.append(
            {
                "diff_type": "only_base",
                "key_column": key_label,
                "key_value": key,
                "column": "",
                "base_value": _row_to_string(base_rows[key]),
                "compare_value": "",
            }
        )

    for key in sorted(compare_keys - base_keys):
        diff_rows.append(
            {
                "diff_type": "only_compare",
                "key_column": key_label,
                "key_value": key,
                "column": "",
                "base_value": "",
                "compare_value": _row_to_string(compare_rows[key]),
            }
        )

    stats = {
        "base_rows": len(base_rows),
        "compare_rows": len(compare_rows),
        "matched": len(matched_keys),
        "changed_rows": changed,
        "only_base": len(base_keys - compare_keys),
        "only_compare": len(compare_keys - base_keys),
        "identical": len(matched_keys) - changed,
    }
    return diff_rows, stats


def _resolve_value_columns(
    base_headers: List[str],
    compare_headers: List[str],
    base_key: str,
    compare_key: str,
    requested: Optional[List[str]],
) -> Tuple[List[str], List[str]]:
    """确定值列列表（以基准表列名为准，需两表都存在），返回 (值列, 被跳过的列)。"""
    base_cols = [c for c in base_headers if c != base_key]
    if requested:
        selected = [c for c in requested if c != base_key]
        missing = [c for c in selected if c not in base_headers]
        skipped = [c for c in selected if c in base_headers and c not in compare_headers]
        skipped += [f"{c}（基准表不存在）" for c in missing]
        cols = [c for c in selected if c in base_headers and c in compare_headers]
        return cols, skipped
    cols = [c for c in base_cols if c in compare_headers and c != compare_key]
    skipped = [c for c in base_cols if c not in compare_headers]
    return cols, skipped


def _now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ---------------------------------------------------------------- 输出：差异明细表

def build_diff_workbook(diff_rows: List[Dict[str, Any]], stats: Dict[str, int]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "差异明细"

    headers = ["差异类型", "键列名", "键值", "涉及列", "基准值", "比对值"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    if diff_rows:
        for row in diff_rows:
            ws.append(
                [
                    TYPE_LABELS[row["diff_type"]],
                    row["key_column"],
                    row["key_value"],
                    row["column"],
                    row["base_value"],
                    row["compare_value"],
                ]
            )
            ws.cell(row=ws.max_row, column=1).fill = FILL_DIFF_TYPE[row["diff_type"]]
    else:
        ws.cell(row=2, column=1, value="两个表格完全一致，未发现差异")

    # 统计 sheet
    ws2 = wb.create_sheet("统计")
    stat_items = [
        ("基准表行数", stats["base_rows"]),
        ("比对表行数", stats["compare_rows"]),
        ("键匹配数", stats["matched"]),
        ("值差异行数", stats["changed_rows"]),
        ("仅基准有", stats["only_base"]),
        ("仅比对有", stats["only_compare"]),
        ("完全一致", stats["identical"]),
    ]
    for name, value in stat_items:
        ws2.append([name, value])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for sheet in (ws, ws2):
        for col_cells in sheet.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(width + 4, 12), 60)
        sheet.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------- 输出：高亮

def build_highlight_workbook(
    target_path: Path,
    sheet_name: str,
    key_col: str,
    rows_other: Dict[str, Dict[str, str]],
    diff_rows: List[Dict[str, Any]],
    target_side: str,
) -> bytes:
    """以 target 文件为底版生成高亮副本。

    值差异单元格标黄并加批注（基准值→比对值）；底版独有行整行着色；
    另一侧独有行追加到数据末尾并整行着色。rows_other 是另一侧行数据，供追加行取值。
    """
    wb = load_workbook(target_path)
    if sheet_name not in wb.sheetnames:
        raise ExcelDiffError(f"工作簿中不存在 sheet「{sheet_name}」")
    ws = wb[sheet_name]
    headers = ["" if c.value is None else str(c.value) for c in ws[1]]
    if key_col not in headers:
        raise ExcelDiffError(f"表中不存在键列「{key_col}」")
    key_idx = headers.index(key_col)

    # key -> 行号
    key_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_idx + 1).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        key_to_row[normalize_cell(v)] = r

    changed_by_key: Dict[str, List[Dict[str, Any]]] = {}
    only_target: List[str] = []
    only_other: List[str] = []
    for d in diff_rows:
        if d["diff_type"] == "changed":
            changed_by_key.setdefault(d["key_value"], []).append(d)
        elif d["diff_type"] == "only_base":
            (only_target if target_side == "base" else only_other).append(d["key_value"])
        else:  # only_compare
            (only_target if target_side == "compare" else only_other).append(d["key_value"])

    # 1) 值差异单元格：黄色 + 批注
    for key, items in changed_by_key.items():
        row_idx = key_to_row.get(key)
        if row_idx is None:
            continue
        for item in items:
            col = item["column"]
            if col not in headers:
                continue
            cell = ws.cell(row=row_idx, column=headers.index(col) + 1)
            cell.fill = FILL_CHANGED
            cell.comment = Comment(
                f"基准值：{item['base_value']}\n比对值：{item['compare_value']}", "Lab Tools"
            )

    # 2) 底版独有行整行着色
    fill_own = FILL_ONLY_BASE if target_side == "base" else FILL_ONLY_COMPARE
    for key in only_target:
        row_idx = key_to_row.get(key)
        if row_idx is None:
            continue
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if cell.value is not None:
                cell.fill = fill_own

    # 3) 另一侧独有行追加到末尾
    if only_other:
        fill_other = FILL_ONLY_COMPARE if target_side == "base" else FILL_ONLY_BASE
        next_row = ws.max_row + 1
        for key in only_other:
            values = rows_other.get(key, {})
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=next_row, column=c)
                cell.value = values.get(h, "")
                cell.fill = fill_other
            next_row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------- 入口

def run_compare(
    base_path: Path,
    compare_path: Path,
    base_name: str,
    compare_name: str,
    req: CompareRequest,
) -> Dict[str, Any]:
    """执行比对，返回 {bytes, download_name, output_mode, stats, skipped_columns, preview, preview_total}。"""
    output_mode = req.output_mode
    if output_mode not in ("diff", "highlight"):
        raise ExcelDiffError("输出模式仅支持 diff（差异明细表）或 highlight（高亮）")
    if output_mode == "highlight" and req.highlight_target not in ("base", "compare"):
        raise ExcelDiffError("高亮底版仅支持 base（基准表）或 compare（比对表）")

    # 解析两个 sheet
    if output_mode == "highlight":
        base_rows, _, base_headers = rows_from_sheet(
            _load_sheet_ws(base_path, req.base_sheet), req.base_key_column
        )
        compare_rows, _, compare_headers = rows_from_sheet(
            _load_sheet_ws(compare_path, req.compare_sheet), req.compare_key_column
        )
    else:
        base_rows, base_headers = _rows_diff(base_path, req.base_sheet, req.base_key_column)
        compare_rows, compare_headers = _rows_diff(compare_path, req.compare_sheet, req.compare_key_column)

    value_cols, skipped = _resolve_value_columns(
        base_headers, compare_headers, req.base_key_column, req.compare_key_column, req.value_columns
    )
    if not value_cols:
        raise ExcelDiffError("没有可用于比对的列：请检查键列设置，或两表至少有一列公共表头")

    diff_rows, stats = compute_diff(base_rows, compare_rows, req.base_key_column, value_cols)

    if output_mode == "diff":
        content = build_diff_workbook(diff_rows, stats)
        download_name = f"差异明细_{_now_tag()}.xlsx"
    else:
        target_side = req.highlight_target
        target_path, target_sheet, target_key = (
            (base_path, req.base_sheet, req.base_key_column)
            if target_side == "base"
            else (compare_path, req.compare_sheet, req.compare_key_column)
        )
        rows_other = compare_rows if target_side == "base" else base_rows
        content = build_highlight_workbook(
            target_path, target_sheet, target_key, rows_other, diff_rows, target_side
        )
        stem = Path(base_name if target_side == "base" else compare_name).stem
        download_name = f"{stem}_高亮对比.xlsx"

    return {
        "bytes": content,
        "download_name": download_name,
        "output_mode": output_mode,
        "stats": stats,
        "skipped_columns": skipped,
        "preview": diff_rows[:PREVIEW_LIMIT],
        "preview_total": len(diff_rows),
    }


def _rows_diff(path: Path, sheet_name: str, key_col: str):
    sheets = load_sheets(path)
    if sheet_name not in sheets:
        raise ExcelDiffError(f"工作簿中不存在 sheet「{sheet_name}」")
    df = sheets[sheet_name]
    return rows_from_df(df, key_col)


def _load_sheet_ws(path: Path, sheet_name: str):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ExcelDiffError(f"工作簿中不存在 sheet「{sheet_name}」")
    return wb[sheet_name]
