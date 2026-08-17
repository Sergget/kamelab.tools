import io

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


def make_workbook_bytes(dfs: dict) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=name)
    return buffer.getvalue()


BASE_DF = pd.DataFrame(
    {
        "工号": ["A01", "A02", "A03", "A04"],
        "姓名": ["张三", "李四", "王五", "赵六"],
        "工资": [100, 200, 300, 400],
        "部门": ["销售", "研发", "研发", "测试"],
    }
)

COMPARE_DF = pd.DataFrame(
    {
        "工号": ["A01", "A02", "A03", "A05"],
        "姓名": ["张三", "李四", "王五改", "钱七"],
        "工资": [150, 200, 300, 500],
        "部门": ["销售", "研发", "研发", "测试"],
    }
)


def upload(content: bytes, fname: str) -> dict:
    response = client.post(
        "/api/tools/excel-diff/upload",
        files={"file": (fname, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    return response.json()


@pytest.fixture(scope="module")
def files():
    base = upload(make_workbook_bytes({"员工表": BASE_DF}), "基准表.xlsx")
    comp = upload(make_workbook_bytes({"员工表": COMPARE_DF}), "比对表.xlsx")
    return base, comp


def test_tools_registry():
    tools = client.get("/api/tools").json()
    ids = [t["id"] for t in tools]
    assert "excel-diff" in ids


def test_upload(files):
    base, comp = files
    assert base["file_name"] == "基准表.xlsx"
    assert base["sheets"][0]["name"] == "员工表"
    assert base["sheets"][0]["rows"] == 4
    assert "工号" in base["sheets"][0]["headers"]


def test_compare_diff_mode(files):
    base, comp = files
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": base["file_id"],
            "compare_file_id": comp["file_id"],
            "base_sheet": "员工表",
            "compare_sheet": "员工表",
            "base_key_column": "工号",
            "compare_key_column": "工号",
            "output_mode": "diff",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    stats = data["stats"]
    # A01 工资 100->150；A02 一致；A03 姓名 王五->王五改；A04 仅基准；A05 仅比对
    assert stats["base_rows"] == 4
    assert stats["compare_rows"] == 4
    assert stats["matched"] == 3
    assert stats["changed_rows"] == 2
    assert stats["only_base"] == 1
    assert stats["only_compare"] == 1
    assert stats["identical"] == 1

    types = {p["diff_type"] for p in data["preview"]}
    assert {"changed", "only_base", "only_compare"} <= types
    assert data["preview_total"] == 4  # 2 值差异 + 1 仅基准 + 1 仅比对

    # 下载产物校验
    dl = client.get(f"/api/tools/excel-diff/download/{data['job_id']}")
    assert dl.status_code == 200
    wb = load_workbook(io.BytesIO(dl.content))
    assert "差异明细" in wb.sheetnames
    assert "统计" in wb.sheetnames
    ws = wb["差异明细"]
    header = [c.value for c in ws[1]]
    assert header == ["差异类型", "键列名", "键值", "涉及列", "基准值", "比对值"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 4


def test_compare_select_value_columns(files):
    base, comp = files
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": base["file_id"],
            "compare_file_id": comp["file_id"],
            "base_sheet": "员工表",
            "compare_sheet": "员工表",
            "base_key_column": "工号",
            "compare_key_column": "工号",
            "value_columns": ["工资"],
            "output_mode": "diff",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    # 只比工资列：A01 工资不同、A03 姓名变化不纳入
    assert data["stats"]["changed_rows"] == 1
    cols = {p["column"] for p in data["preview"] if p["diff_type"] == "changed"}
    assert cols == {"工资"}


def test_compare_highlight_base(files):
    base, comp = files
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": base["file_id"],
            "compare_file_id": comp["file_id"],
            "base_sheet": "员工表",
            "compare_sheet": "员工表",
            "base_key_column": "工号",
            "compare_key_column": "工号",
            "output_mode": "highlight",
            "highlight_target": "base",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["download_name"].endswith("_高亮对比.xlsx")
    assert data["output_mode"] == "highlight"

    dl = client.get(f"/api/tools/excel-diff/download/{data['job_id']}")
    assert dl.status_code == 200
    wb = load_workbook(io.BytesIO(dl.content))
    ws = wb["员工表"]
    # 表头第 1 行：工号=1, 姓名=2, 工资=3, 部门=4
    # A01 在数据第 2 行，工资列(3)应为黄色
    assert ws.cell(row=2, column=3).fill.start_color.rgb.endswith("FFFF00")
    # A03 姓名列(2) 第 4 行黄色
    assert ws.cell(row=4, column=2).fill.start_color.rgb.endswith("FFFF00")
    # A04（仅基准）第 5 行整行橙色
    assert ws.cell(row=5, column=1).fill.start_color.rgb.endswith("F4B084")
    # A05（仅比对）追加到最后一行（第 6 行）浅绿
    assert ws.max_row >= 6
    assert ws.cell(row=6, column=1).fill.start_color.rgb.endswith("A9D08E")
    # A01 工资单元格带批注
    assert ws.cell(row=2, column=3).comment is not None
    assert "100" in ws.cell(row=2, column=3).comment.text


def test_compare_highlight_compare(files):
    base, comp = files
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": base["file_id"],
            "compare_file_id": comp["file_id"],
            "base_sheet": "员工表",
            "compare_sheet": "员工表",
            "base_key_column": "工号",
            "compare_key_column": "工号",
            "output_mode": "highlight",
            "highlight_target": "compare",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    dl = client.get(f"/api/tools/excel-diff/download/{data['job_id']}")
    wb = load_workbook(io.BytesIO(dl.content))
    ws = wb["员工表"]
    # 底版为比对表：A01 工资(150)黄色；A05（仅比对）整行绿色；A04（仅基准）追加橙色
    assert ws.cell(row=2, column=3).fill.start_color.rgb.endswith("FFFF00")
    assert ws.cell(row=5, column=1).fill.start_color.rgb.endswith("A9D08E")  # A05 在比对表第 5 行
    assert ws.cell(row=6, column=1).fill.start_color.rgb.endswith("F4B084")  # A04 追加


def test_compare_identical():
    df = pd.DataFrame({"id": ["1", "2"], "v": [10, 20]})
    f1 = upload(make_workbook_bytes({"s": df}), "a.xlsx")
    f2 = upload(make_workbook_bytes({"s": df.copy()}), "b.xlsx")
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": f1["file_id"],
            "compare_file_id": f2["file_id"],
            "base_sheet": "s",
            "compare_sheet": "s",
            "base_key_column": "id",
            "compare_key_column": "id",
            "output_mode": "diff",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["stats"]["identical"] == 2
    assert data["stats"]["changed_rows"] == 0
    assert data["preview_total"] == 0


def test_compare_duplicate_key():
    df1 = pd.DataFrame({"id": ["1", "1"], "v": [10, 20]})
    df2 = pd.DataFrame({"id": ["1"], "v": [10]})
    f1 = upload(make_workbook_bytes({"s": df1}), "a.xlsx")
    f2 = upload(make_workbook_bytes({"s": df2}), "b.xlsx")
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": f1["file_id"],
            "compare_file_id": f2["file_id"],
            "base_sheet": "s",
            "compare_sheet": "s",
            "base_key_column": "id",
            "compare_key_column": "id",
            "output_mode": "diff",
        },
    )
    assert resp.status_code == 400
    assert "重复值" in resp.json()["detail"]


def test_compare_missing_key_column():
    df1 = pd.DataFrame({"id": ["1"], "v": [10]})
    df2 = pd.DataFrame({"id": ["1"], "v": [10]})
    f1 = upload(make_workbook_bytes({"s": df1}), "a.xlsx")
    f2 = upload(make_workbook_bytes({"s": df2}), "b.xlsx")
    resp = client.post(
        "/api/tools/excel-diff/compare",
        json={
            "base_file_id": f1["file_id"],
            "compare_file_id": f2["file_id"],
            "base_sheet": "s",
            "compare_sheet": "s",
            "base_key_column": "不存在列",
            "compare_key_column": "id",
            "output_mode": "diff",
        },
    )
    assert resp.status_code == 400
    assert "不存在键列" in resp.json()["detail"]


def test_upload_rejects_bad_ext():
    resp = client.post(
        "/api/tools/excel-diff/upload",
        files={"file": ("a.txt", b"hello", "text/plain")},
    )
    assert resp.status_code == 400


def test_upload_corrupted_xlsx():
    resp = client.post(
        "/api/tools/excel-diff/upload",
        files={"file": ("bad.xlsx", b"not a real xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "无法解析" in resp.json()["detail"]
